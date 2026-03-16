#!/usr/bin/env python3
"""
读取附件五Excel文件，比较施工内容和单项结算金额单价，并根据比较结果进行颜色标记。
"""

import pandas as pd
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill

# 定义文件路径
attachment_file = "附件五：中国铁塔云南公司2024-2025年土建杆塔施工集中采购项目施工模块安全生产费明细表.xlsx"
generated_file = "外包结算单测试模版.xlsx"

def read_attachment_five():
    """读取附件五Excel文件，提取模块名称和红河/文山列数据"""
    print(f"读取附件五文件: {attachment_file}")
    
    if not os.path.exists(attachment_file):
        print(f"错误: 文件 {attachment_file} 不存在")
        return None
    
    # 读取Excel文件
    xl = pd.ExcelFile(attachment_file)
    print(f"工作表: {xl.sheet_names}")
    
    # 读取"土建+杆塔"工作表
    if '土建+杆塔' in xl.sheet_names:
        df = pd.read_excel(attachment_file, sheet_name='土建+杆塔')
        print(f"读取工作表: 土建+杆塔")
    else:
        df = pd.read_excel(attachment_file, sheet_name=xl.sheet_names[0])
        print(f"读取工作表: {xl.sheet_names[0]}")
    
    print(f"数据形状: {df.shape}")
    print(f"列名: {df.columns.tolist()}")
    
    # 查找模块名称和红河/文山列
    module_col = None
    price_col = None
    
    for col in df.columns:
        col_str = str(col).strip()
        if '模块名称' in col_str:
            module_col = col
            print(f"找到模块名称列: {col}")
        if '红河' in col_str or '文山' in col_str:
            price_col = col
            print(f"找到红河/文山列: {col}")
    
    if module_col is None or price_col is None:
        print("错误: 未找到模块名称或红河/文山列")
        return None
    
    # 提取数据
    data = df[[module_col, price_col]].copy()
    data.columns = ['施工内容', '单项结算金额单价']
    
    # 删除空值行和标题行
    data = data.dropna(subset=['施工内容'])
    
    # 删除包含"模块名称"的行（标题行）
    data = data[data['施工内容'] != '模块名称']
    
    print(f"\n附件五提取的数据（前10行）:")
    print(data.head(10))
    print(f"\n总共 {len(data)} 条数据")
    
    return data

def read_generated_file():
    """读取生成的Excel文件，提取施工内容和单项结算金额单价"""
    print(f"\n读取生成的文件: {generated_file}")
    
    if not os.path.exists(generated_file):
        print(f"错误: 文件 {generated_file} 不存在")
        return None
    
    # 读取Excel文件
    xl = pd.ExcelFile(generated_file)
    print(f"工作表: {xl.sheet_names}")
    
    # 读取工程结算金额工作表
    if '工程结算金额' in xl.sheet_names:
        df = pd.read_excel(generated_file, sheet_name='工程结算金额')
    else:
        df = pd.read_excel(generated_file, sheet_name=xl.sheet_names[0])
    
    print(f"数据形状: {df.shape}")
    print(f"列名: {df.columns.tolist()}")
    
    # 查找施工内容和单项结算金额单价列
    content_col = None
    price_col = None
    
    for col in df.columns:
        col_str = str(col).strip()
        if '施工内容' in col_str:
            content_col = col
            print(f"找到施工内容列: {col}")
        if '单项结算金额单价' in col_str:
            price_col = col
            print(f"找到单项结算金额单价列: {col}")
    
    if content_col is None or price_col is None:
        print("错误: 未找到施工内容或单项结算金额单价列")
        return None
    
    # 提取数据
    data = df[[content_col, price_col]].copy()
    data.columns = ['施工内容', '单项结算金额单价']
    
    print(f"\n生成文件提取的数据:")
    print(data)
    
    return data

def compare_data(attachment_data, generated_data):
    """比较两组数据，返回比较结果"""
    print("\n开始比较数据...")
    
    # 创建结果数据框
    result_data = generated_data.copy()
    result_data['比较结果'] = ''
    result_data['附件五单价'] = None  # 添加附件五的单价列用于对比，初始化为None
    
    # 将附件五数据转换为字典，方便查找
    attachment_dict = {}
    for _, row in attachment_data.iterrows():
        content = str(row['施工内容']).strip()
        price = row['单项结算金额单价']
        attachment_dict[content] = price
    
    print(f"附件五数据字典（前5条）: {dict(list(attachment_dict.items())[:5])}")
    
    # 比较每一行
    for idx, row in result_data.iterrows():
        content = str(row['施工内容']).strip()
        price = row['单项结算金额单价']
        
        if content in attachment_dict:
            attachment_price = attachment_dict[content]
            result_data.at[idx, '附件五单价'] = attachment_price
            
            # 比较价格是否相同
            try:
                # 转换为浮点数进行比较
                gen_price = float(price) if pd.notna(price) else 0
                att_price = float(attachment_price) if pd.notna(attachment_price) else 0
                
                # 允许一定的误差范围（0.01）
                if abs(gen_price - att_price) < 0.01:
                    result_data.at[idx, '比较结果'] = '相同'
                else:
                    result_data.at[idx, '比较结果'] = '不一致'
            except (ValueError, TypeError):
                # 如果无法转换为数字，直接比较字符串
                if str(price) == str(attachment_price):
                    result_data.at[idx, '比较结果'] = '相同'
                else:
                    result_data.at[idx, '比较结果'] = '不一致'
        else:
            result_data.at[idx, '比较结果'] = '不存在'
            result_data.at[idx, '附件五单价'] = None
    
    print(f"\n比较结果统计:")
    print(result_data['比较结果'].value_counts())
    
    return result_data

def add_missing_data(result_data, attachment_data):
    """将附件五中存在但生成文件中不存在的数据添加到结果末尾"""
    print("\n检查缺失数据...")
    
    # 获取生成文件中已有的施工内容
    existing_contents = set(str(content).strip() for content in result_data['施工内容'])
    
    # 查找缺失的数据
    missing_data = []
    for _, row in attachment_data.iterrows():
        content = str(row['施工内容']).strip()
        if content not in existing_contents:
            missing_data.append({
                '施工内容': content,
                '单项结算金额单价': row['单项结算金额单价'],
                '附件五单价': row['单项结算金额单价'],
                '比较结果': '不存在'
            })
    
    if missing_data:
        print(f"发现 {len(missing_data)} 条缺失数据，添加到结果末尾")
        missing_df = pd.DataFrame(missing_data)
        result_data = pd.concat([result_data, missing_df], ignore_index=True)
    else:
        print("没有缺失数据")
    
    return result_data

def save_with_colors(result_data, output_file):
    """保存结果到Excel文件，并根据比较结果添加颜色标记"""
    print(f"\n保存结果到文件: {output_file}")
    
    # 首先使用pandas保存为普通Excel文件
    result_data.to_excel(output_file, index=False, engine='xlsxwriter')
    
    # 然后使用openpyxl打开并添加颜色
    from openpyxl import load_workbook
    
    wb = load_workbook(output_file)
    ws = wb.active
    
    # 定义颜色
    green_fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')  # 浅绿色
    red_fill = PatternFill(start_color='FFB6C1', end_color='FFB6C1', fill_type='solid')    # 浅红色
    yellow_fill = PatternFill(start_color='FFFFE0', end_color='FFFFE0', fill_type='solid')  # 浅黄色
    
    # 找到单项结算金额单价列的索引
    price_col_idx = None
    for idx, cell in enumerate(ws[1], 1):
        if cell.value == '单项结算金额单价':
            price_col_idx = idx
            break
    
    if price_col_idx is None:
        print("错误: 未找到单项结算金额单价列")
        return
    
    # 找到比较结果列的索引
    result_col_idx = None
    for idx, cell in enumerate(ws[1], 1):
        if cell.value == '比较结果':
            result_col_idx = idx
            break
    
    if result_col_idx is None:
        print("错误: 未找到比较结果列")
        return
    
    # 根据比较结果添加颜色
    for row in ws.iter_rows(min_row=2):
        result_cell = row[result_col_idx - 1]
        price_cell = row[price_col_idx - 1]
        
        if result_cell.value == '相同':
            price_cell.fill = green_fill
        elif result_cell.value == '不一致':
            price_cell.fill = red_fill
        elif result_cell.value == '不存在':
            price_cell.fill = yellow_fill
    
    # 保存文件
    wb.save(output_file)
    print(f"成功保存带颜色标记的文件: {output_file}")

def main():
    """主函数"""
    print("=" * 60)
    print("开始比较结算数据")
    print("=" * 60)
    
    # 1. 读取附件五数据
    attachment_data = read_attachment_five()
    if attachment_data is None:
        print("无法读取附件五数据，程序退出")
        return
    
    # 2. 读取生成的文件
    generated_data = read_generated_file()
    if generated_data is None:
        print("无法读取生成文件数据，程序退出")
        return
    
    # 3. 比较数据
    result_data = compare_data(attachment_data, generated_data)
    
    # 4. 添加缺失数据
    result_data = add_missing_data(result_data, attachment_data)
    
    # 5. 保存带颜色标记的结果
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_file = f"结算比较结果_{timestamp}.xlsx"
    save_with_colors(result_data, output_file)
    
    print("\n" + "=" * 60)
    print("比较完成!")
    print(f"结果文件: {output_file}")
    print("=" * 60)

if __name__ == "__main__":
    main()
